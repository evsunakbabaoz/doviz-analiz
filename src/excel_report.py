"""
excel_report.py
---------------
Excel raporu üreten modül. (Hocanın en çok önem verdiği kısım!)
Görev: Farklı sekmelere (sheets) sahip zengin bir Excel dosyası oluştur.

Sekmeler:
  1. Data       - Ham veriler
  2. Summary    - İstatistiksel özet
  3. Weekly     - Haftalık ortalamalar
  4. Analysis   - Araştırma sorusu cevabı
"""

import pandas as pd
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.series import DataPoint
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("UYARI: openpyxl kurulu değil. 'pip install openpyxl' komutunu çalıştırın.")

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")


def create_excel_report(df, stats_df, weekly_df, stable_currency, stable_vol,
                        best_currency, best_pct):
    """
    Ana Excel raporu oluşturma fonksiyonu.
    Parametreler:
        df           : Temizlenmiş ham veri
        stats_df     : İstatistik özeti
        weekly_df    : Haftalık özet
        stable_currency: En stabil döviz adı
        stable_vol   : En stabil dövizin volatilitesi
        best_currency: En çok değer kazanan döviz
        best_pct     : Değer kazanma yüzdesi
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)

    today = datetime.today().strftime("%Y%m%d")
    filename = f"doviz_raporu_{today}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    print(f"Excel raporu oluşturuluyor: {filename}")

    # ExcelWriter ile birden fazla sekme yaz
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # --- SEKME 1: Data (Ham Veri) ---
        df_export = df.copy()
        df_export["tarih"] = df_export["tarih"].dt.strftime("%Y-%m-%d")
        df_export.to_excel(writer, sheet_name="Data", index=False)
        print("  Sekme 1 yazıldı: Data")

        # --- SEKME 2: Summary (İstatistiksel Özet) ---
        stats_df.to_excel(writer, sheet_name="Summary", index=False)
        print("  Sekme 2 yazıldı: Summary")

        # --- SEKME 3: Weekly (Haftalık Ortalamalar) ---
        weekly_df.to_excel(writer, sheet_name="Weekly", index=False)
        print("  Sekme 3 yazıldı: Weekly")

        # --- SEKME 4: Analysis (Araştırma Sorusu Cevabı) ---
        analysis_data = {
            "Araştırma Sorusu": [
                "Hangi döviz TL karşısında en stabil seyretti?",
                "Son dönemde en çok değer kazanan döviz hangisi?",
                "",
                "SONUÇLAR",
                "En Stabil Döviz",
                "Volatilite (%)",
                "",
                "En Çok Değer Kazanan",
                "Toplam Değişim (%)",
                "",
                "Rapor Tarihi"
            ],
            "Cevap": [
                "",
                "",
                "",
                "",
                stable_currency,
                f"{stable_vol:.4f}%",
                "",
                best_currency,
                f"{best_pct:+.2f}%",
                "",
                datetime.today().strftime("%d/%m/%Y %H:%M")
            ]
        }
        analysis_df = pd.DataFrame(analysis_data)
        analysis_df.to_excel(writer, sheet_name="Analysis", index=False)
        print("  Sekme 4 yazıldı: Analysis")

    # Openpyxl ile biçimlendirme uygula
    if OPENPYXL_AVAILABLE:
        _format_excel(filepath)
        print(f"  Biçimlendirme uygulandı.")

    print(f"Excel raporu hazır: {filepath}")
    return filepath


def _format_excel(filepath):
    """
    Oluşturulan Excel dosyasına profesyonel biçimlendirme uygular.
    Başlık renkleri, koyu yazı, sütun genişlikleri vb.
    """
    wb = load_workbook(filepath)

    # Renk paleti
    HEADER_BLUE = "1F4E79"    # Koyu mavi başlık
    HEADER_GREEN = "1E6B3C"   # Koyu yeşil başlık
    HEADER_ORANGE = "833C00"  # Turuncu başlık
    HEADER_PURPLE = "4B0082"  # Mor başlık
    LIGHT_BLUE = "D6E4F0"     # Açık mavi satır
    LIGHT_GREEN = "D5F5E3"    # Açık yeşil satır

    sheet_colors = {
        "Data": HEADER_BLUE,
        "Summary": HEADER_GREEN,
        "Weekly": HEADER_ORANGE,
        "Analysis": HEADER_PURPLE,
    }

    for sheet_name, header_color in sheet_colors.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Başlık satırı biçimlendirmesi
        header_fill = PatternFill(start_color=header_color,
                                  end_color=header_color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Satır yüksekliğini ayarla
        ws.row_dimensions[1].height = 30

        # Veri satırları: Zebra şerit (almaşık renkler)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_fill_color = LIGHT_BLUE if row_idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=row_fill_color,
                                   end_color=row_fill_color, fill_type="solid")
            for cell in row:
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        # Sütun genişliklerini otomatik ayarla
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 4, 40)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Freeze panes (başlık satırını sabitle)
        ws.freeze_panes = "A2"

    wb.save(filepath)


# --- Test ---
if __name__ == "__main__":
    import numpy as np
    import sys
    sys.path.append(os.path.dirname(__file__))
    from analyzer import calculate_statistics, calculate_weekly_summary

    print("=== excel_report.py TEST ===")

    from datetime import datetime, timedelta
    records = []
    base_date = datetime(2024, 1, 1)
    for i in range(30):
        for currency, base in [("USD", 32), ("EUR", 35), ("GBP", 41)]:
            records.append({
                "tarih": base_date + timedelta(days=i),
                "kur": round(base * (1 + np.random.uniform(-0.02, 0.02)), 4),
                "para_birimi": currency,
                "aykiri_deger": False
            })

    df = pd.DataFrame(records)
    df["tarih"] = pd.to_datetime(df["tarih"])

    stats = calculate_statistics(df)
    weekly = calculate_weekly_summary(df)

    path = create_excel_report(
        df=df,
        stats_df=stats,
        weekly_df=weekly,
        stable_currency="EUR",
        stable_vol=0.0012,
        best_currency="USD",
        best_pct=3.45
    )
    print(f"\nRapor oluşturuldu: {path}")
